import os, re, glob, json, time, argparse, logging, hashlib
from datetime import datetime, date
import pdfplumber
import pandas as pd
from tqdm import tqdm
from openpyxl.utils import get_column_letter

# ── CONFIGS ───────────────────────────────────────────────────────────────────
# LEMBRETE: no GitHub Actions use caminhos RELATIVOS
from pathlib import Path
ROOT = Path(".")
PDF_DIR        = str(ROOT / "inbox")                       # onde os PDFs baixam
OUT_DIR        = ROOT / "out"
OUT_DIR.mkdir(parents=True, exist_ok=True)

MATRIX_XLSX    = str(OUT_DIR / "OFERTASMATRIZ.xlsx")
PARQUET_OFS    = str(OUT_DIR / "OFERTASMATRIZ_OFERTAS.parquet")
PARQUET_ERR    = str(OUT_DIR / "OFERTASMATRIZ_ERROS.parquet")

SHEET_OFERTAS  = "OFERTAS"
SHEET_ERROS    = "ERRO_MONITORAMENTO"

# Arquivos auxiliares (joguei pro out/)
ROW_IDS_FILE   = str(OUT_DIR / "OFERTASMATRIZ_ROW_IDS.txt")
ERR_IDS_FILE   = str(OUT_DIR / "OFERTASMATRIZ_ERR_IDS.txt")
STATE_JSON     = str(OUT_DIR / "OFERTASMATRIZ_STATE.json")

# [NOVO] Parquet “final”
REPLACE_OFERTAS_PATH = str(OUT_DIR / "OFERTAS.parquet")

# Loop do script original não será usado no Actions (vamos rodar --once)
LOOP_INTERVAL_SEC = 10 * 60
# ──────────────────────────────────────────────────────────────────────────────
logging.getLogger("pdfminer").setLevel(logging.ERROR)

VALID_ENTITIES = {
    "123milhas","agoda","airbnb","aircanada","airfrance","aeromexico","americanairlines",
    "ancoradouro","azul","booking.com","capoviagens","cestarollitravel","confianca",
    "cvc","decolar","esferatur","expedia","flipmilhas","flytourgapnet","gol",
    "googleflights","gotogate","hoteis.com","hurb","iberia","jetsmart","kayak",
    "kissandfly","kiwi.com","latam","lufthansa","maxmilhas","momondo","mrsmrssmith",
    "mytrip","passagenspromo","primetour","queensberryviagens","rexturadvance",
    "sakuratur","skyscanner","submarinoviagens","tap","trendoperadora","traveloka",
    "trip.com","unitedairlines","viajanet","visualturismo","voepass","vrbo","zarpo","zupper"
}
AIRLINES = ["gol","latam","azul","voepass","jetsmart","airfrance","unitedairlines",
            "iberia","lufthansa","aeromexico","aircanada","americanairlines","tap"]

# Regras
dates_regex   = re.compile(r"(\d{2}/\d{2}/\d{4},\s*\d{2}:\d{2})")
price_regex   = re.compile(r"R\$\s*([\d\s\.,]+)")
CUTOFF_OFFERS = "complemente sua viagem"
TIMES_CUTOFF  = "verificando preços e disponibilidade"
FIRST_PAGE_ERROR_RULES = {
    re.compile(r"as melhores ofertas e promoções", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"destinos nacionais mais buscados", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"encaminhando para o website soli", re.IGNORECASE):  "ERRO DE PAGINA",
    re.compile(r"passagens aéreas em promoção\s*\|\s*l", re.IGNORECASE): "ERRO DE PAGINA",
    re.compile(r"skyscanner\s+você\s+é\s+uma\s+pessoa\s+ou", re.IGNORECASE): "ERRO ANTIBOT",
}
PAGE_ERROR_PATTERNS = {
    re.compile(r"passagens aéreas.*hotéis.*aluguel de carros", re.IGNORECASE): "ErroPaginaInicial",
    re.compile(r"pacotes de viagens", re.IGNORECASE): "ErroPaginaDecolar"
}
MONTH_MAP = {"jan":1,"fev":2,"mar":3,"abr":4,"mai":5,"jun":6,"jul":7,"ago":8,"set":9,"out":10,"nov":11,"dez":12}

# ── Helpers de normalização/ID (hash) ─────────────────────────────────────────
OF_ID_COLS = [
    "Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
    "Tipo de Voo","Data do Voo","Data/Hora da Busca","Agência/Companhia",
    "Preço","TRECHO","ADVP"   # Ranking de fora
]
ER_ID_COLS = ["Nome do Arquivo","Erro","Trecho","Pagina"]

def to_upper_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return df
    return df.applymap(lambda x: x.upper() if isinstance(x, str) else x)

def _fmt_date_series(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, dayfirst=True, errors="coerce").dt.strftime("%d/%m/%Y").fillna("")

def _fmt_price_series(s: pd.Series) -> pd.Series:
    s2 = pd.to_numeric(s, errors="coerce").round(2)
    return s2.map(lambda x: f"{x:.2f}" if pd.notna(x) else "").astype(str)

def _canon_ofertas(df: pd.DataFrame) -> pd.DataFrame:
    c = df.copy()
    for col in ["Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
                "Tipo de Voo","Agência/Companhia","TRECHO"]:
        if col in c: c[col] = c[col].astype(str).str.strip().str.upper()
    if "Data do Voo" in c:            c["Data do Voo"] = _fmt_date_series(c["Data do Voo"])
    if "Data/Hora da Busca" in c:
        so_data = c["Data/Hora da Busca"].astype(str).str.extract(r"(\d{2}/\d{2}/\d{4})", expand=False)
        c["Data/Hora da Busca"] = _fmt_date_series(so_data)
    if "Preço" in c:                  c["Preço"] = _fmt_price_series(c["Preço"])
    if "ADVP" in c:                   c["ADVP"] = pd.to_numeric(c["ADVP"], errors="coerce").fillna(0).astype(int).astype(str)
    for col in OF_ID_COLS:
        if col not in c: c[col] = ""
        c[col] = c[col].fillna("")
    return c

def _canon_erros(df: pd.DataFrame) -> pd.DataFrame:
    c = df.copy()
    for col in ["Nome do Arquivo","Erro","Trecho"]:
        if col in c: c[col] = c[col].astype(str).str.strip().str.upper()
    if "Pagina" in c: c["Pagina"] = pd.to_numeric(c["Pagina"], errors="coerce").fillna(1).astype(int).astype(str)
    for col in ER_ID_COLS:
        if col not in c: c[col] = ""
        c[col] = c[col].fillna("")
    return c

def _hash_concat(df: pd.DataFrame, cols: list) -> pd.Series:
    s = df[cols[0]].astype(str)
    for col in cols[1:]:
        s = s.str.cat(df[col].astype(str), sep="||")
    return s.map(lambda t: hashlib.sha1(t.encode("utf-8")).hexdigest())

def build_row_ids(df: pd.DataFrame) -> pd.Series: return _hash_concat(_canon_ofertas(df), OF_ID_COLS)
def build_err_ids(df: pd.DataFrame) -> pd.Series: return _hash_concat(_canon_erros(df), ER_ID_COLS)

# (REMOVIDO) Funções de cache/estado — mantidas apenas como no código original, porém não usadas.
def load_id_cache(path: str) -> set:
    return set()

def append_id_cache(path: str, ids: pd.Series):
    return

# [INSERÇÃO] Cache simples de PDFs por mtime+tamanho ---------------------------
def load_state() -> dict:
    """Lê JSON de estado de PDFs processados (caminho -> assinatura)."""
    try:
        with open(STATE_JSON, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_state(state: dict):
    """Grava JSON de estado."""
    try:
        os.makedirs(os.path.dirname(STATE_JSON), exist_ok=True)
        with open(STATE_JSON, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
# -----------------------------------------------------------------------------

# ── Extração PDF ──────────────────────────────────────────────────────────────
def first_page_error_code(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages: return None, None
            text = (pdf.pages[0].extract_text() or "")
    except Exception:
        return None, None
    low = text.lower()
    for pat, code in FIRST_PAGE_ERROR_RULES.items():
        m = pat.search(low)
        if m:
            start = max(0, m.start()-25); end = min(len(text), m.end()+25)
            trecho = text[start:end].replace("\n"," ").strip()
            return code, trecho
    return None, None

def extract_flight_info(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ""
    low = text.lower()

    for pat, err in PAGE_ERROR_PATTERNS.items():
        if pat.search(text): return None, err

    idx = low.find(TIMES_CUTOFF)
    snippet = text[:idx] if idx != -1 else text

    all_times = re.findall(r"\b(\d{2}:\d{2})\b", snippet)
    times_dict = {f"Horário{i+1}": t for i, t in enumerate(all_times)}

    tipo = ""
    matches = list(re.finditer(r"\b(\d{2}:\d{2})\b", snippet))
    if len(matches) >= 2:
        pos2 = matches[1].end()
        window = snippet[pos2:pos2+200].lower()
        if re.search(r"\bdireto\b", window): tipo = "DIRETO"
        else:
            m_esc = re.search(r"(\d+)\s*escalas?", window)
            m_par = re.search(r"(\d+)\s*paradas?", window)
            if m_esc: tipo = f"{m_esc.group(1)} ESCALAS"
            elif m_par: tipo = f"{m_par.group(1)} PARADAS"

    cia = next((c.upper() for c in AIRLINES if c in low), "")

    dm = re.search(r"ida[^\d]*(\d{1,2})\s+de\s+([a-zç]+)\.?\s+de\s+(\d{4})", low)
    if dm:
        day, m_pt, yr = dm.groups()
        m = MONTH_MAP.get(m_pt[:3], 0)
        flight_date = f"{int(day):02d}/{m:02d}/{yr}"
    else:
        flight_date = ""

    return {"Companhia Aérea": cia, **times_dict, "Tipo de Voo": tipo, "Data do Voo": flight_date}, None

def extract_offers_from_pdf(pdf_path, search_dt):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for p in pdf.pages:
            text += (p.extract_text() or "") + "\n"

    text = re.sub(r"(\d)\n(\d)", r"\1\2", text)
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    cutoff = next((i for i,l in enumerate(lines) if CUTOFF_OFFERS in l.lower()), len(lines))
    lines = lines[:cutoff]

    if not search_dt:
        for l in lines[:10]:
            m = dates_regex.search(l)
            if m: search_dt = m.group(1); break

    offers, last_ent = [], None
    for l in lines:
        norm = re.sub(r"\s+","", l.lower())
        for ent in VALID_ENTITIES:
            if ent.replace(" ","") in norm:
                last_ent = ent; break
        pm = price_regex.search(l)
        if pm and last_ent:
            raw = pm.group(1)
            num = re.sub(r"[^\d,]","", raw).replace(",", ".")
            try: price = float(num)
            except: continue
            offers.append({"Agência/Companhia": last_ent, "Preço": price})
            last_ent = None
    return offers, search_dt

def get_trecho(file_name):
    parts = file_name[:6].upper().split('_')
    return f"{parts[0]}-{parts[1]}" if len(parts)>1 else parts[0]

def rank_prices(df):
    df['Ranking'] = df.groupby('Nome do Arquivo')['Preço'].rank(method='min', ascending=True).astype(int)
    return df

def todas_colunas_preenchidas(row, cols_req):
    return all(pd.notna(row.get(col)) and str(row.get(col)).strip() != "" for col in cols_req)

# ── Excel/Parquet ─────────────────────────────────────────────────────────────
def write_back_preserving(file_path, df_ofertas, df_erros):
    mode = "a" if os.path.exists(file_path) else "w"
    with pd.ExcelWriter(file_path, engine="openpyxl", mode=mode, if_sheet_exists=("replace" if mode=="a" else None)) as writer:
        df_ofertas.to_excel(writer, index=False, sheet_name=SHEET_OFERTAS)
        df_erros.to_excel(writer, index=False, sheet_name=SHEET_ERROS)
        # formatar datas como DD/MM/YYYY na planilha OFERTAS
        ws = writer.sheets[SHEET_OFERTAS]
        cols = list(df_ofertas.columns)
        for col_name in ["Data do Voo","Data/Hora da Busca"]:
            if col_name in cols:
                idx = cols.index(col_name) + 1
                letter = get_column_letter(idx)
                for cell in ws[letter][1:]:
                    cell.number_format = "DD/MM/YYYY"

# [NOVO] escrita segura do arquivo OFERTAS.parquet pedido pela Tassi
def _replace_ofertas_parquet_safely(ofertas_df: pd.DataFrame):
    try:
        os.makedirs(os.path.dirname(REPLACE_OFERTAS_PATH), exist_ok=True)
        tmp_path = REPLACE_OFERTAS_PATH + ".tmp"
        of = ofertas_df.copy()
        for c in ["Data do Voo","Data/Hora da Busca"]:
            if c in of: of[c] = pd.to_datetime(of[c], errors="coerce")
        of.to_parquet(tmp_path, index=False)
        os.replace(tmp_path, REPLACE_OFERTAS_PATH)
        print(f"💾 Substituído: {REPLACE_OFERTAS_PATH} | Linhas: {len(of)}")
    except Exception as e:
        print(f"⚠️ Falha ao substituir {REPLACE_OFERTAS_PATH}: {e}")

def export_parquet(ofertas_df, erros_df):
    # Converte datas pro tipo datetime64[ns] no Parquet
    of = ofertas_df.copy()
    for c in ["Data do Voo","Data/Hora da Busca"]:
        if c in of: of[c] = pd.to_datetime(of[c], errors="coerce")
    of.to_parquet(PARQUET_OFS, index=False)      # requer pyarrow
    if erros_df is not None and not erros_df.empty:
        er = erros_df.copy()
        er.to_parquet(PARQUET_ERR, index=False)

    # [NOVO] sempre substituir o OFERTAS.parquet final pedido
    if of is not None:
        _replace_ofertas_parquet_safely(of)

# ── 1 ciclo de atualização ────────────────────────────────────────────────────
def run_cycle():
    # (REMOVIDO) state/cache; processa TODOS os PDFs sempre
    pdfs = sorted(glob.glob(os.path.join(PDF_DIR, "*.pdf")))
    print(f"[{datetime.now():%H:%M:%S}] PDFs na pasta: {len(pdfs)} | Processando todos (sem cache)")
    if not pdfs:
        return None, None, 0, 0

    # [INSERÇÃO] Seleciona apenas PDFs novos/alterados (cache por mtime+tamanho)
    state = load_state()
    def _sig(p: str) -> str | None:
        try:
            st = os.stat(p)
            return f"{st.st_size}-{int(st.st_mtime)}"
        except Exception:
            return None
    to_process = []
    for p in pdfs:
        sig = _sig(p)
        if sig is None:
            continue
        if state.get(p) != sig:
            to_process.append(p)
    if not to_process:
        print(f"[{datetime.now():%H:%M:%S}] Nada novo. Todos os PDFs já convertidos.")
        return pd.DataFrame(), pd.DataFrame(), 0, 0
    print(f"[{datetime.now():%H:%M:%S}] Novos/alterados: {len(to_process)} de {len(pdfs)}")

    offers_rows, errors_rows = [], []

    for path in tqdm(to_process, desc="Processando PDFs"):   # [INSERÇÃO] usa to_process
        fn = os.path.basename(path)

        code, trecho = first_page_error_code(path)
        if code:
            errors_rows.append({"Nome do Arquivo": fn, "Erro": code, "Trecho": trecho, "Pagina": 1})

        flight_info, err = extract_flight_info(path)
        if err:
            errors_rows.append({"Nome do Arquivo": fn, "Erro": err, "Trecho": "", "Pagina": 1})

        offers, sdt = extract_offers_from_pdf(path, "")
        for o in offers:
            offers_rows.append({
                "Nome do Arquivo": fn,
                **(flight_info or {}),
                "Data/Hora da Busca": sdt,  # ex: "09/08/2025, 13:13"
                **o,
                "TRECHO": get_trecho(fn)
            })

    # DataFrames novos
    new_offers_df = pd.DataFrame(offers_rows)
    new_erros_df  = pd.DataFrame(errors_rows)

    # ── OFERTAS: datas, ADVP (subtração direta), ranking, filtros, maiúsculo ──
    if not new_offers_df.empty:
        new_offers_df["Data do Voo"] = pd.to_datetime(
            new_offers_df["Data do Voo"], dayfirst=True, errors="coerce"
        )
        so_data = new_offers_df["Data/Hora da Busca"].astype(str).str.extract(
            r"(\d{2}/\d{2}/\d{4})", expand=False
        )
        new_offers_df["Data/Hora da Busca"] = pd.to_datetime(
            so_data, dayfirst=True, errors="coerce"
        )

        diff_days = (
            new_offers_df["Data do Voo"].dt.normalize()
            - new_offers_df["Data/Hora da Busca"].dt.normalize()
        ).dt.days
        new_offers_df["ADVP"] = diff_days.fillna(0).astype(int)

        new_offers_df = rank_prices(new_offers_df)

        req = ["Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
               "Tipo de Voo","Data do Voo","Data/Hora da Busca",
               "Agência/Companhia","Preço","TRECHO","ADVP","Ranking"]
        new_offers_df = new_offers_df[new_offers_df.apply(lambda r: todas_colunas_preenchidas(r, req), axis=1)]
        new_offers_df = new_offers_df[new_offers_df["Agência/Companhia"].str.lower() != "skyscanner"]

        new_offers_df = to_upper_df(new_offers_df)

    # ── ERROS: maiúsculo ──
    if not new_erros_df.empty:
        new_erros_df = to_upper_df(new_erros_df)

    # Carrega base atual
    try: base_ofertas = pd.read_excel(MATRIX_XLSX, sheet_name=SHEET_OFERTAS, engine="openpyxl")
    except: base_ofertas = pd.DataFrame()
    try: base_erros = pd.read_excel(MATRIX_XLSX, sheet_name=SHEET_ERROS, engine="openpyxl")
    except: base_erros = pd.DataFrame()

    base_cols = list(base_ofertas.columns) if not base_ofertas.empty else \
        ["Nome do Arquivo","Companhia Aérea","Horário1","Horário2","Horário3",
         "Tipo de Voo","Data do Voo","Data/Hora da Busca","Agência/Companhia",
         "Preço","TRECHO","ADVP","Ranking"]
    err_cols = list(base_erros.columns) if not base_erros.empty else ["Nome do Arquivo","Erro","Trecho","Pagina","EM_AMBAS"]

    novos_unicos = new_offers_df.copy() if not new_offers_df.empty else pd.DataFrame()
    new_errs_unique = new_erros_df.copy() if not new_erros_df.empty else pd.DataFrame()

    if not novos_unicos.empty:
        for c in base_cols:
            if c not in novos_unicos.columns: novos_unicos[c] = pd.NA
        novos_unicos = novos_unicos[base_cols]
    if not new_errs_unique.empty:
        if "EM_AMBAS" not in new_errs_unique.columns: new_errs_unique["EM_AMBAS"] = pd.NA
        for c in err_cols:
            if c not in new_errs_unique.columns: new_errs_unique[c] = pd.NA
        new_errs_unique = new_errs_unique[err_cols]

    final_ofertas = pd.concat([base_ofertas, novos_unicos], ignore_index=True) if not novos_unicos.empty else base_ofertas
    final_erros   = pd.concat([base_erros, new_errs_unique], ignore_index=True) if not new_errs_unique.empty else base_erros

    if not final_erros.empty:
        ofertas_set = set(final_ofertas["Nome do Arquivo"].dropna().astype(str)) if "Nome do Arquivo" in final_ofertas else set()
        final_erros["EM_AMBAS"] = final_erros["Nome do Arquivo"].apply(lambda x: "SIM" if str(x) in ofertas_set else "NÃO")

    # Grava Excel + Parquet (sempre que há algo novo OU base vazia)
    if not novos_unicos.empty or not new_errs_unique.empty or base_ofertas.empty or base_erros.empty:
        write_back_preserving(MATRIX_XLSX, final_ofertas, final_erros)
        export_parquet(final_ofertas, final_erros)

    # [INSERÇÃO] Atualiza o cache de PDFs processados
    for p in to_process:
        try:
            st = os.stat(p)
            state[p] = f"{st.st_size}-{int(st.st_mtime)}"
        except Exception:
            pass
    save_state(state)

    return final_ofertas, final_erros, len(novos_unicos), len(new_errs_unique)

# ── CLI / Execução ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--once", action="store_true", help="Executa apenas 1 ciclo.")
    args = ap.parse_args()

    if args.once:
        of_df, er_df, n_of, n_er = run_cycle()
        print(f"✅ Ciclo concluído. Ofertas novas: {n_of} | Erros novos: {n_er}")
    else:
        # [NOVO] Loop de 10 em 10 minutos
        while True:
            try:
                start = datetime.now()
                of_df, er_df, n_of, n_er = run_cycle()
                print(f"✅ Ciclo concluído {start:%d/%m %H:%M:%S}. Ofertas novas: {n_of} | Erros novos: {n_er}")
            except Exception as e:
                print(f"❌ Erro no ciclo: {e}")
            print(f"⏲️ Próxima execução em 10 minutos ({(datetime.now() + pd.Timedelta(seconds=LOOP_INTERVAL_SEC)):%H:%M:%S})")
            time.sleep(LOOP_INTERVAL_SEC)

# =====================================================================
# convert_ofertasmatriz_to_parquet.py  (mantido como no original)
# Lê OFERTASMATRIZ.* na pasta dada, pega a aba "OFERTAS" (se existir)
# e salva como OFERTAS.parquet no mesmo diretório.
# =====================================================================

import glob
import sys
import json  # [INSERÇÃO]

BASE_DIR = r"C:\Users\tassiana.silva\Downloads\teste"
BASENAME = "OFERTASMATRIZ"          # vamos procurar OFERTASMATRIZ.xlsx/.xlsm/.xlsb/.xls
OUT_NAME = "OFERTAS.parquet"        # nome do arquivo parquet de saída
ALLOWED_EXT = (".xlsx", ".xlsm", ".xlsb", ".xls")

# [INSERÇÃO] Cache do conversor (evita reconverter arquivo idêntico)
CONV_STATE = os.path.join(BASE_DIR, "OFERTAS_CONVERT_STATE.json")

def _load_conv_state() -> dict:   # [INSERÇÃO]
    try:
        with open(CONV_STATE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_conv_state(state: dict):  # [INSERÇÃO]
    try:
        with open(CONV_STATE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _sig(path: str) -> str | None:  # [INSERÇÃO]
    try:
        st = os.stat(path)
        return f"{st.st_size}-{int(st.st_mtime)}"
    except Exception:
        return None

def find_input_file(base_dir: str, basename: str) -> str | None:
    # tenta pelos formatos mais comuns
    for ext in ALLOWED_EXT:
        path = os.path.join(base_dir, basename + ext)
        if os.path.exists(path):
            return path
    # fallback: qualquer extensão que exista com esse basename
    hits = sorted(glob.glob(os.path.join(base_dir, basename + ".*")))
    return hits[0] if hits else None

def read_ofertas_sheet(xls_path: str) -> pd.DataFrame:
    # Abre o Excel e usa a aba "OFERTAS" se existir; senão, a primeira aba
    try:
        xl = pd.ExcelFile(xls_path)
    except Exception as e:
        print(f"❌ Não consegui abrir '{xls_path}': {e}")
        sys.exit(1)

    # Procura por 'OFERTAS' ignorando maiúsc/minúsc
    oferta_sheet = None
    for s in xl.sheet_names:
        if s.strip().upper() == "OFERTAS":
            oferta_sheet = s
            break
    sheet_to_use = oferta_sheet if oferta_sheet else xl.sheet_names[0]

    print(f"→ Lendo aba: {sheet_to_use}")
    df = xl.parse(sheet_to_use)
    return df

def main():
    in_file = find_input_file(BASE_DIR, BASENAME)
    if not in_file:
        print(f"❌ Arquivo '{BASENAME}.*' não encontrado em {BASE_DIR}.")
        sys.exit(1)

    print(f"→ Arquivo encontrado: {in_file}")

    # [INSERÇÃO] Verifica cache: só converte se o arquivo mudou
    state = _load_conv_state()
    in_sig = _sig(in_file)
    out_path = os.path.join(BASE_DIR, OUT_NAME)
    if in_sig and state.get("input_sig") == in_sig and os.path.exists(out_path):
        print(f"⏭️  Nada novo para converter. Fonte inalterada e '{OUT_NAME}' já existe.")
        print(f"   Fonte: {in_file} | Sinal: {in_sig}")
        print(f"   Saída: {out_path}")
        return

    df = read_ofertas_sheet(in_file)

    # Conversão de tipos amigável ao Parquet (opcional)
    df = df.convert_dtypes()

    try:
        df.to_parquet(out_path, engine="pyarrow", compression="snappy", index=False)
    except Exception as e:
        print("❌ Falhou ao salvar Parquet. Verifique se 'pyarrow' está instalado (pip install pyarrow).")
        print(f"Detalhes: {e}")
        sys.exit(1)

    # [INSERÇÃO] Atualiza cache do conversor
    state["input_sig"] = in_sig
    state["output_path"] = out_path
    _save_conv_state(state)

    print(f"✅ Salvo: {out_path}  | Linhas: {len(df)}")

if __name__ == "__main__":
    main()
